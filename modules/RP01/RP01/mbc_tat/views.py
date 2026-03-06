from flask import render_template, request, jsonify, session, redirect, url_for, Response
from functools import wraps
from datetime import datetime, date
import io
import traceback

from .. import bp
from database import get_db, get_cursor


# ── auth ───────────────────────────────────────────────────────────────────────
def login_required(f):
    @wraps(f)
    def decorated(*args, **kwargs):
        if 'user_id' not in session:
            return redirect(url_for('login'))
        return f(*args, **kwargs)
    return decorated


# ── time helpers ───────────────────────────────────────────────────────────────
def _parse(ts):
    """Parse a timestamp value to datetime, handling datetime, date, and string inputs."""
    if ts is None:
        return None
    if isinstance(ts, datetime):
        return ts
    # datetime.date but not datetime.datetime
    if isinstance(ts, date):
        return datetime(ts.year, ts.month, ts.day, 0, 0, 0)
    s = str(ts).strip()
    if not s:
        return None
    for fmt in ('%Y-%m-%d %H:%M:%S', '%Y-%m-%d %H:%M', '%Y-%m-%dT%H:%M:%S', '%Y-%m-%dT%H:%M', '%Y-%m-%d'):
        try:
            return datetime.strptime(s, fmt)
        except ValueError:
            continue
    return None


def _diff_mins(ts_from, ts_to):
    """Duration in minutes, or None if either timestamp is missing/negative."""
    a, b = _parse(ts_from), _parse(ts_to)
    if not a or not b:
        return None
    delta = (b - a).total_seconds()
    return delta / 60 if delta >= 0 else None


def _avg_mins(values):
    """Average of non-None values; None if list is empty."""
    valid = [v for v in values if v is not None]
    return sum(valid) / len(valid) if valid else None


def _fmt_mins(total_mins):
    """Format float minutes as H:MM, or '—' if None."""
    if total_mins is None:
        return '—'
    h, m = divmod(int(round(total_mins)), 60)
    return f'{h}:{m:02d}'


def _fy_range(d):
    """Return (fy_start_str, fy_end_str) ISO strings for the FY containing date d."""
    y, m = d.year, d.month
    if m >= 4:
        return f'{y}-04-01', f'{y + 1}-03-31'
    return f'{y - 1}-04-01', f'{y}-03-31'


def _fy_label(d):
    y, m = d.year, d.month
    if m >= 4:
        return f'FY {str(y)[2:]}-{str(y + 1)[2:]}'
    return f'FY {str(y - 1)[2:]}-{str(y)[2:]}'


# ── data fetch ─────────────────────────────────────────────────────────────────
def _fetch_trips(from_date_str, to_date_str):
    """Return raw rows for Import MBC trips within [from_date, to_date] (ISO strings)."""
    conn = get_db()
    cur  = get_cursor(conn)
    cur.execute("""
        SELECT
            lp.arrived_load_port,    lp.loading_commenced,   lp.loading_completed,
            lp.cast_off_load_port,
            dp.arrival_gull_island,  dp.departure_gull_island, dp.vessel_arrival_port,
            dp.unloading_commenced,  dp.unloading_completed,
            dp.vessel_cast_off,      dp.sailed_out_load_port
        FROM mbc_header h
        LEFT JOIN mbc_load_port_lines      lp ON lp.mbc_id = h.id
        LEFT JOIN mbc_discharge_port_lines dp ON dp.mbc_id = h.id
        WHERE h.operation_type = 'Import'
          AND h.doc_date >= %s
          AND h.doc_date <= %s
        ORDER BY h.doc_date ASC, h.id ASC
    """, (from_date_str, to_date_str))
    rows = cur.fetchall()
    conn.close()
    # Convert RealDictRow → plain dict so the connection can safely be closed
    return [dict(r) for r in rows]


# ── metrics ────────────────────────────────────────────────────────────────────
def _compute_metrics(rows):
    """Compute per-segment average durations (minutes) across all trips."""
    buckets = {
        'preberthing':       [],
        'loading':           [],
        'wait_after_load':   [],
        'total_jaigad':      [],
        'jaigad_to_gull':    [],
        'gull_waiting':      [],
        'gull_to_dhar':      [],
        'jaigad_to_dhar':    [],
        'preberthing_dhar':  [],
        'unloading':         [],
        'wait_after_unload': [],
        'total_dharamtar':   [],
        'dhar_to_jaigad':    [],
        'tat':               [],
    }

    for r in rows:
        al = r.get('arrived_load_port')
        lc = r.get('loading_commenced')
        lp = r.get('loading_completed')
        co = r.get('cast_off_load_port')
        ag = r.get('arrival_gull_island')
        dg = r.get('departure_gull_island')
        ad = r.get('vessel_arrival_port')
        uc = r.get('unloading_commenced')
        up = r.get('unloading_completed')
        cd = r.get('vessel_cast_off')
        so = r.get('sailed_out_load_port')

        buckets['preberthing'].append(_diff_mins(al, lc))
        buckets['loading'].append(_diff_mins(lc, lp))
        buckets['wait_after_load'].append(_diff_mins(lp, co))
        buckets['total_jaigad'].append(_diff_mins(al, co))
        buckets['jaigad_to_gull'].append(_diff_mins(co, ag))
        buckets['gull_waiting'].append(_diff_mins(ag, dg))
        buckets['gull_to_dhar'].append(_diff_mins(dg, ad))
        buckets['jaigad_to_dhar'].append(_diff_mins(co, ad))
        buckets['preberthing_dhar'].append(_diff_mins(ad, uc))
        buckets['unloading'].append(_diff_mins(uc, up))
        buckets['wait_after_unload'].append(_diff_mins(up, cd))
        buckets['total_dharamtar'].append(_diff_mins(ad, cd))
        buckets['dhar_to_jaigad'].append(_diff_mins(cd, so))
        buckets['tat'].append(_diff_mins(al, so))

    return {k: _avg_mins(v) for k, v in buckets.items()}


# ── report builder ─────────────────────────────────────────────────────────────
def _build_report(sel_date_str):
    try:
        sel = datetime.strptime(sel_date_str, '%Y-%m-%d').date()
    except (ValueError, TypeError):
        sel = date.today()

    sel_iso = sel.strftime('%Y-%m-%d')

    # ── date period ──────────────────────────────────────────────────────
    date_rows    = _fetch_trips(sel_iso, sel_iso)
    date_count   = len(date_rows)
    date_metrics = _compute_metrics(date_rows)
    date_label   = sel.strftime('%d-%m-%Y')

    # ── MTD period ──────────────────────────────────────────────────────
    mtd_start_iso = sel.replace(day=1).strftime('%Y-%m-%d')
    mtd_rows      = _fetch_trips(mtd_start_iso, sel_iso)
    mtd_count     = len(mtd_rows)
    mtd_metrics   = _compute_metrics(mtd_rows)
    mtd_label     = sel.strftime('%b-%y')

    # ── YTD period ──────────────────────────────────────────────────────
    fy_start_iso, _  = _fy_range(sel)
    ytd_rows         = _fetch_trips(fy_start_iso, sel_iso)
    ytd_count        = len(ytd_rows)
    ytd_metrics      = _compute_metrics(ytd_rows)
    ytd_label        = _fy_label(sel)

    def _row(label, key, style='data'):
        return {
            'label':    label,
            'style':    style,
            'date_val': _fmt_mins(date_metrics.get(key)),
            'mtd_val':  _fmt_mins(mtd_metrics.get(key)),
            'ytd_val':  _fmt_mins(ytd_metrics.get(key)),
        }

    def _total_row(label, key, style):
        return {
            'label':    label,
            'style':    style,
            'date_val': _fmt_mins(date_metrics.get(key)),
            'mtd_val':  _fmt_mins(mtd_metrics.get(key)),
            'ytd_val':  _fmt_mins(ytd_metrics.get(key)),
        }

    rows = [
        # ── Jaigad section ───────────────────────────────────────────────
        _row('Jaigad Arrival -  Jaigad Loading Commenced  (Preberthing delay)', 'preberthing'),
        _row('Loading Commence - Loading Completion  (Loading time)',            'loading'),
        _row('Loading Completed - Cast Off from Jaigad  (Waiting after loading)', 'wait_after_load'),
        _total_row('Total time taken at Jaigad', 'total_jaigad', 'section_total'),

        # ── Transit section ─────────────────────────────────────────────
        _row('Jaigad Departure to Gull Arrival  (Loaded Transit time)', 'jaigad_to_gull'),
        _row('Gull Arrival - Gull Departure  (Waiting at Gull)',         'gull_waiting'),
        _row('Gull Departure - Dharamatar Arrival',                      'gull_to_dhar'),
        _total_row('Jaigad Departure - Dharamtar Arrival (Jaigad to Dharamtar)', 'jaigad_to_dhar', 'main_total'),

        # ── Dharamtar section ────────────────────────────────────────────
        _row('Dharamtar Arrival to Disch Commenced  (Preberthing delay)', 'preberthing_dhar'),
        _row('Disch Commended to Disch Completed  (Unloading Time)',       'unloading'),
        _row('Disch Completed to Cast Off from Dharamtar  (Waiting after Unloading)', 'wait_after_unload'),
        _total_row('Total time taken at Dharamtar', 'total_dharamtar', 'section_total'),

        # ── Return + TAT ─────────────────────────────────────────────────
        _total_row('Dharamtar Departure to Jaigad Arrival', 'dhar_to_jaigad', 'main_total'),
        _total_row('TAT', 'tat', 'tat'),
    ]

    return {
        'date_label':  date_label,
        'mtd_label':   mtd_label,
        'ytd_label':   ytd_label,
        'date_trips':  date_count,
        'mtd_trips':   mtd_count,
        'ytd_trips':   ytd_count,
        'rows':        rows,
    }


# ── routes ─────────────────────────────────────────────────────────────────────
@bp.route('/module/RP01/mbc-tat/')
@login_required
def mbc_tat_index():
    return render_template('mbc_tat/mbc_tat_report.html', username=session.get('username'))


@bp.route('/api/module/RP01/mbc-tat/data')
@login_required
def mbc_tat_data():
    sel_date = request.args.get('date', date.today().strftime('%Y-%m-%d'))
    try:
        return jsonify(_build_report(sel_date))
    except Exception:
        return jsonify({'error': traceback.format_exc()}), 500


@bp.route('/api/module/RP01/mbc-tat/download')
@login_required
def mbc_tat_download():
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

    sel_date = request.args.get('date', date.today().strftime('%Y-%m-%d'))
    try:
        data = _build_report(sel_date)
    except Exception:
        return jsonify({'error': traceback.format_exc()}), 500

    wb = Workbook()
    ws = wb.active
    ws.title = 'MBC TAT Report'

    # ── colour palette ───────────────────────────────────────────────────────
    GREEN_HDR  = 'CCFF99'
    YELLOW     = 'FFFF00'
    ORANGE     = 'FFCC00'
    CYAN       = '00FFFF'
    WHITE      = 'FFFFFF'
    RED_TXT    = 'FF0000'
    ORANGE_TXT = 'FF6600'

    thin = Side(style='thin', color='000000')
    bdr  = Border(left=thin, right=thin, top=thin, bottom=thin)
    ctr  = Alignment(horizontal='center', vertical='center', wrap_text=True)
    left = Alignment(horizontal='left',   vertical='center', wrap_text=True)

    def fill(hex_color):
        return PatternFill('solid', fgColor=hex_color)

    def font(bold=False, color='000000', size=11):
        return Font(name='Calibri', bold=bold, color=color, size=size)

    # ── Row 1: period label headers ──────────────────────────────────────────
    ws.row_dimensions[1].height = 20
    ws.cell(1, 1, '').border = bdr

    c2 = ws.cell(1, 2, data['date_label'])
    c2.font = Font(name='Calibri', bold=True, color=RED_TXT, size=11)
    c2.fill, c2.border, c2.alignment = fill(GREEN_HDR), bdr, ctr

    c3 = ws.cell(1, 3, 'MTD')
    c3.font = Font(name='Calibri', bold=True, color=RED_TXT, size=11)
    c3.fill, c3.border, c3.alignment = fill(GREEN_HDR), bdr, ctr

    c4 = ws.cell(1, 4, 'YTD')
    c4.font = Font(name='Calibri', bold=True, color=ORANGE_TXT, size=11)
    c4.fill, c4.border, c4.alignment = fill(GREEN_HDR), bdr, ctr

    # ── Row 2: column headers ────────────────────────────────────────────────
    ws.row_dimensions[2].height = 24
    for ci, val in enumerate(['Activity', data['date_label'], data['mtd_label'], data['ytd_label']], 1):
        c = ws.cell(2, ci, val)
        c.font   = font(bold=True)
        c.fill   = fill(GREEN_HDR)
        c.border = bdr
        c.alignment = ctr

    # ── Row 3: Trips ─────────────────────────────────────────────────────────
    ws.row_dimensions[3].height = 18
    for ci, val in enumerate(['Trips', data['date_trips'], data['mtd_trips'], data['ytd_trips']], 1):
        c = ws.cell(3, ci, val)
        c.font   = font(bold=True)
        c.fill   = fill(YELLOW)
        c.border = bdr
        c.alignment = ctr

    # ── Data rows ─────────────────────────────────────────────────────────────
    STYLE_FILL = {
        'data':          WHITE,
        'section_total': YELLOW,
        'main_total':    ORANGE,
        'tat':           CYAN,
    }
    STYLE_BOLD = {'data': False, 'section_total': True, 'main_total': True, 'tat': True}

    for ri, row in enumerate(data['rows'], start=4):
        ws.row_dimensions[ri].height = 18
        style   = row['style']
        bg      = STYLE_FILL.get(style, WHITE)
        is_bold = STYLE_BOLD.get(style, False)

        for ci, (val, aln) in enumerate([
            (row['label'],    left),
            (row['date_val'], ctr),
            (row['mtd_val'],  ctr),
            (row['ytd_val'],  ctr),
        ], 1):
            c = ws.cell(ri, ci, val)
            c.font      = font(bold=is_bold)
            c.fill      = fill(bg)
            c.border    = bdr
            c.alignment = aln

    # ── Footnotes ─────────────────────────────────────────────────────────────
    fn_row = len(data['rows']) + 4
    footnotes = [
        '*Unloading time includes the shifting time after partial discharge of MBC & '
        'discharge kept on hold to give priority for other MBC.',
        '*Preberthing time includes the MBC waiting at Berth 10 for the unloading turn.',
    ]
    for i, fn in enumerate(footnotes):
        r = fn_row + i
        ws.row_dimensions[r].height = 30
        ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=4)
        c = ws.cell(r, 1, fn)
        c.font      = Font(name='Calibri', italic=True, size=9, color='404040')
        c.alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
        c.border    = bdr

    # ── Column widths ─────────────────────────────────────────────────────────
    ws.column_dimensions['A'].width = 60
    ws.column_dimensions['B'].width = 14
    ws.column_dimensions['C'].width = 14
    ws.column_dimensions['D'].width = 12

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)

    fname = f'MBC_TAT_Report_{sel_date}.xlsx'
    return Response(
        buf.getvalue(),
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        headers={'Content-Disposition': f'attachment; filename="{fname}"'},
    )
