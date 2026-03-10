from flask import render_template, request, session, redirect, url_for, Response
from functools import wraps
from datetime import date, datetime, timedelta
import io

from .. import bp
from database import get_db, get_cursor

from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# ── Style constants (same as rest of RP01) ──────────────────────────────────
XL_NORM_SZ = 11
_thin  = Side(style='thin',   color='000000')
_bdr   = Border(left=_thin, right=_thin, top=_thin, bottom=_thin)
_ctr   = Alignment(horizontal='center', vertical='center', wrap_text=False)
_left  = Alignment(horizontal='left',   vertical='center', wrap_text=False)

# Hardcoded annual totals by fiscal year — stored for reference, not printed in report
ANNUAL_ROUTE_TOTALS = {
    "2024-2025": {},
    "2025-2026": {},
}


def _fill(hex_color):
    return PatternFill('solid', fgColor=hex_color)


def _font(bold=False, size=XL_NORM_SZ):
    return Font(name='Calibri', bold=bold, size=size)


def _parse_dt(val):
    if not val:
        return None
    if isinstance(val, datetime):
        return val
    try:
        return datetime.fromisoformat(str(val))
    except Exception:
        return None


def _fmt_dt(val, strfmt='%d-%m-%Y %H:%M'):
    dt = _parse_dt(val)
    return dt.strftime(strfmt) if dt else ''


def login_required(f):
    @wraps(f)
    def decorated(*args, **kwargs):
        if 'user_id' not in session:
            return redirect(url_for('login'))
        return f(*args, **kwargs)
    return decorated


@bp.route('/module/RP01/daily-ops/')
@login_required
def daily_ops_index():
    return render_template('daily_ops/daily_ops.html', username=session.get('username'))


def _fetch_data(report_date, operation_type):
    window_end   = datetime(report_date.year, report_date.month, report_date.day, 7, 0, 0)
    window_start = window_end - timedelta(hours=24)
    ws_str = window_start.strftime('%Y-%m-%d %H:%M:%S')
    we_str = window_end.strftime('%Y-%m-%d %H:%M:%S')

    conn = get_db()
    cur  = get_cursor(conn)

    # Active vessels with NOR present, filtered by operation type and 7am window
    cur.execute("""
        SELECT h.id, h.vcn_id, h.vessel_name, h.operation_type,
               h.nor_tendered, h.discharge_commenced, h.discharge_completed
        FROM ldud_header h
        WHERE h.discharge_commenced IS NOT NULL
          AND h.nor_tendered IS NOT NULL AND h.nor_tendered != ''
          AND h.operation_type = %s
          AND h.discharge_commenced < %s
          AND (h.discharge_completed IS NULL OR h.discharge_completed > %s)
        ORDER BY h.discharge_commenced ASC
        LIMIT 5
    """, (operation_type, we_str, ws_str))
    vessels = [dict(r) for r in cur.fetchall()]

    ldud_ids = [v['id']     for v in vessels]
    vcn_ids  = [v['vcn_id'] for v in vessels if v.get('vcn_id')]

    bl_import = {}
    bl_export  = {}
    vcn_meta   = {}
    if vcn_ids:
        cur.execute("""
            SELECT vcn_id, COALESCE(SUM(bl_quantity), 0) AS total
            FROM vcn_cargo_declaration WHERE vcn_id = ANY(%s) GROUP BY vcn_id
        """, (vcn_ids,))
        for r in cur.fetchall():
            bl_import[r['vcn_id']] = float(r['total'])

        cur.execute("""
            SELECT vcn_id, COALESCE(SUM(bl_quantity), 0) AS total
            FROM vcn_export_cargo_declaration WHERE vcn_id = ANY(%s) GROUP BY vcn_id
        """, (vcn_ids,))
        for r in cur.fetchall():
            bl_export[r['vcn_id']] = float(r['total'])

        cur.execute("""
            SELECT id, importer_exporter_name FROM vcn_header WHERE id = ANY(%s)
        """, (vcn_ids,))
        vcn_meta = {r['id']: r['importer_exporter_name'] or '' for r in cur.fetchall()}

    ops_24h     = {}
    ops_till    = {}
    barge_stats = {}

    if ldud_ids:
        cur.execute("""
            SELECT ldud_id, COALESCE(SUM(quantity), 0) AS qty
            FROM ldud_vessel_operations
            WHERE ldud_id = ANY(%s)
              AND start_time >= %s AND start_time < %s
            GROUP BY ldud_id
        """, (ldud_ids, ws_str, we_str))
        for r in cur.fetchall():
            ops_24h[r['ldud_id']] = float(r['qty'])

        cur.execute("""
            SELECT ldud_id, COALESCE(SUM(quantity), 0) AS qty
            FROM ldud_vessel_operations
            WHERE ldud_id = ANY(%s)
              AND start_time < %s
            GROUP BY ldud_id
        """, (ldud_ids, we_str))
        for r in cur.fetchall():
            ops_till[r['ldud_id']] = float(r['qty'])

        _STATUS_KEYS = (
            'at_jetty', 'waiting_discharge', 'waiting_empty_jetty',
            'at_gull_loaded', 'under_loading', 'waiting_loading', 'in_transit_jetty_to_mv',
        )
        cur.execute("""
            SELECT ldud_id, barge_name, discharge_quantity,
                   along_side_vessel, commenced_loading, completed_loading,
                   cast_off_mv, anchored_gull_island, aweigh_gull_island,
                   amf_at_port, along_side_berth, commence_discharge_berth,
                   completed_discharge_berth, cast_off_berth, cast_off_port
            FROM ldud_barge_lines
            WHERE ldud_id = ANY(%s)
              AND along_side_vessel < %s
              AND (cast_off_port IS NULL OR cast_off_port > %s)
        """, (ldud_ids, we_str, ws_str))
        for r in cur.fetchall():
            lid = r['ldud_id']
            bn  = (r['barge_name'] or '').strip()
            qty = r['discharge_quantity']
            if lid not in barge_stats:
                barge_stats[lid] = {'all': set(), **{k: [] for k in _STATUS_KEYS}}
            if bn:
                barge_stats[lid]['all'].add(bn)
            # Determine current status (most-advanced stage wins)
            if r['cast_off_port']:
                status = 'in_transit_jetty_to_mv'
            elif r['completed_discharge_berth'] and not r['cast_off_berth']:
                status = 'waiting_empty_jetty'
            elif r['along_side_berth'] and not r['commence_discharge_berth']:
                status = 'waiting_discharge'
            elif r['amf_at_port'] and not r['along_side_berth']:
                status = 'at_jetty'
            elif r['anchored_gull_island'] and not r['aweigh_gull_island']:
                status = 'at_gull_loaded'
            elif r['commenced_loading'] and not r['completed_loading']:
                status = 'under_loading'
            elif r['along_side_vessel'] and not r['commenced_loading']:
                status = 'waiting_loading'
            else:
                status = None
            if status and bn:
                # At Jetty and Waiting for Discharge include cargo quantity
                if status in ('at_jetty', 'waiting_discharge') and qty:
                    entry = f'{bn} ({int(round(qty))} MT)'
                else:
                    entry = bn
                barge_stats[lid][status].append(entry)

    conn.close()

    for v in vessels:
        lid        = v['id']
        vid        = v.get('vcn_id')
        op         = v.get('operation_type', '')
        bl_qty     = (bl_export.get(vid, 0) if op == 'Export' else bl_import.get(vid, 0)) if vid else 0
        discharged = ops_till.get(lid, 0)
        bs         = barge_stats.get(lid, {})
        _names     = lambda key: ', '.join(bs.get(key, []))
        v['stevedore_group']        = vcn_meta.get(vid, '') if vid else ''
        v['bl_qty']                 = bl_qty
        v['ops_24h']                = ops_24h.get(lid, 0)
        v['ops_till']               = discharged
        v['balance']                = bl_qty - discharged
        v['num_barges']             = len(bs.get('all', set())) or ''
        v['at_jetty']               = _names('at_jetty')
        v['waiting_discharge']      = _names('waiting_discharge')
        v['waiting_empty_jetty']    = _names('waiting_empty_jetty')
        v['at_gull_loaded']         = _names('at_gull_loaded')
        v['under_loading']          = _names('under_loading')
        v['waiting_loading']        = _names('waiting_loading')
        v['in_transit_jetty_to_mv'] = _names('in_transit_jetty_to_mv')

    return vessels


def _fetch_cargo_handled(report_date, operation_type):
    window_end  = datetime(report_date.year, report_date.month, report_date.day, 7, 0, 0)
    window_start = window_end - timedelta(hours=24)
    month_start  = datetime(report_date.year, report_date.month, 1, 7, 0, 0)
    we_str  = window_end.strftime('%Y-%m-%d %H:%M:%S')
    ws_str  = window_start.strftime('%Y-%m-%d %H:%M:%S')
    mth_str = month_start.strftime('%Y-%m-%d %H:%M:%S')

    conn = get_db()
    cur  = get_cursor(conn)

    def _period(start, end):
        cur.execute("""
            SELECT route_name, COALESCE(SUM(quantity), 0) AS qty
            FROM lueu_lines
            WHERE route_name IS NOT NULL AND route_name != ''
              AND operation_type = %s
              AND start_time >= %s AND start_time < %s
            GROUP BY route_name
            ORDER BY route_name
        """, (operation_type, start, end))
        return [(r['route_name'], float(r['qty'])) for r in cur.fetchall()]

    day_rows   = _period(ws_str,  we_str)
    month_rows = _period(mth_str, we_str)
    conn.close()
    return day_rows, month_rows


def _build_excel(vessels, report_date, operation_type, day_rows=None, month_rows=None):
    from openpyxl import Workbook
    wb = Workbook()
    ws = wb.active
    ws.title = 'Daily Ops'
    day_rows   = day_rows   or []
    month_rows = month_rows or []

    # Column widths — wide enough to avoid wrapping
    col_widths = {1: 30, 2: 35, 3: 35, 4: 35, 5: 35, 6: 35, 7: 10, 8: 32, 9: 22}
    for ci, w in col_widths.items():
        ws.column_dimensions[get_column_letter(ci)].width = w

    def _cell(r, c, val='', bold=False, fill='FFFFFF', align=_ctr):
        cell = ws.cell(r, c, val)
        cell.font      = _font(bold=bold)
        cell.fill      = _fill(fill)
        cell.alignment = align
        cell.border    = _bdr
        return cell

    def _merge_row(r, c1, c2, val='', bold=False, fill='FFFFFF', align=_ctr):
        ws.merge_cells(start_row=r, start_column=c1, end_row=r, end_column=c2)
        for ci in range(c1, c2 + 1):
            b = Border(
                left   = _thin if ci == c1 else None,
                right  = _thin if ci == c2 else None,
                top    = _thin,
                bottom = _thin,
            )
            try:
                cell        = ws.cell(r, ci)
                cell.fill   = _fill(fill)
                cell.border = b
            except AttributeError:
                pass  # MergedCell stub in some openpyxl versions
        anchor           = ws.cell(r, c1)
        anchor.value     = val
        anchor.font      = _font(bold=bold)
        anchor.alignment = align

    def _merge_col(r1, r2, c, val='', bold=False, fill='FFFFFF', align=_ctr):
        """Merge cells vertically in column c from row r1 to r2."""
        ws.merge_cells(start_row=r1, start_column=c, end_row=r2, end_column=c)
        for ri in range(r1, r2 + 1):
            b = Border(
                left   = _thin,
                right  = _thin,
                top    = _thin if ri == r1 else None,
                bottom = _thin if ri == r2 else None,
            )
            try:
                cell        = ws.cell(ri, c)
                cell.fill   = _fill(fill)
                cell.border = b
            except AttributeError:
                pass
        anchor           = ws.cell(r1, c)
        anchor.value     = val
        anchor.font      = _font(bold=bold)
        anchor.alignment = align

    date_str  = f"{report_date.day}.{report_date.month}.{report_date.year}"
    title_str = f'Daily Report of JSW Dharamtar Port Operation : {date_str}'

    # Row 1
    ws.row_dimensions[1].height = 20
    _cell(1, 1, report_date.strftime('%d-%m-%Y'), align=_left)
    _merge_row(1, 2, 7, title_str, align=_ctr)
    _cell(1, 8, 'Doc No. | REV.02 | Issue no. 02', align=_left)
    _cell(1, 9, f'Issue Date: {report_date.strftime("%d-%m-%Y")}', align=_left)

    # Row 2: vessel name headers
    ws.row_dimensions[2].height = 20
    _cell(2, 1, '')
    for i, v in enumerate(vessels):
        _cell(2, 2 + i, f'Vessel {i + 1}: {v["vessel_name"]}', bold=True, align=_ctr)
    for i in range(len(vessels), 5):
        _cell(2, 2 + i, '')
    _cell(2, 7, '')
    _cell(2, 8, '')
    _cell(2, 9, '')

    # Data rows (row 3 onwards)
    label_discharge = 'Discharged /Loaded till Date'
    label_balance   = 'Balance on Board /to Load'
    label_commenced = 'Disch Commenced' if operation_type == 'Import' else 'Loading Commenced'
    label_completed = 'Disch Completed' if operation_type == 'Import' else 'Loading Completed'

    _q = lambda x: int(round(x)) if x else ''
    _n = lambda x: x if x else ''
    # (label, field, formatter, align)  — align defaults to _ctr
    ROWS = [
        ('Stevedore/ Barge Group',          'stevedore_group',          None,       _left),
        ('BL Qty',                          'bl_qty',                   _q,         _ctr),
        ('24 hrs Discharge',                'ops_24h',                  _q,         _ctr),
        (label_discharge,                   'ops_till',                 _q,         _ctr),
        (label_balance,                     'balance',                  _q,         _ctr),
        ('Vsl Arrived/NOR',                 'nor_tendered',             _fmt_dt,    _ctr),
        (label_commenced,                   'discharge_commenced',      _fmt_dt,    _ctr),
        (label_completed,                   'discharge_completed',      _fmt_dt,    _ctr),
        (None, None, None, None),
        ('No of Barges',                    'num_barges',               _n,         _ctr),
        ('At Jetty',                        'at_jetty',                 _n,         _left),
        ('Waiting for Discharge',           'waiting_discharge',        _n,         _left),
        ('Waiting Empty at Jetty',          'waiting_empty_jetty',      _n,         _left),
        ('In transit- MV/Gull to Jetty',    None,                       None,       _left),
        ('At Gull- waiting (Loaded)',        'at_gull_loaded',           _n,         _left),
        ('Under Loading at MV',             'under_loading',            _n,         _left),
        ('Waiting for loading',             'waiting_loading',          _n,         _left),
        ('In transit- from Jetty to MV',    'in_transit_jetty_to_mv',   _n,         _left),
    ]

    for idx, (label, field, formatter, align) in enumerate(ROWS):
        r = 3 + idx
        ws.row_dimensions[r].height = 18

        if label is None:
            for ci in range(1, 10):
                _cell(r, ci, '')
            continue

        _cell(r, 1, label, bold=True, align=_left)
        for i, v in enumerate(vessels):
            raw = v.get(field)
            val = formatter(raw) if (formatter and raw is not None) else (raw or '')
            _cell(r, 2 + i, val, align=align)
        for i in range(len(vessels), 5):
            _cell(r, 2 + i, '')
        _cell(r, 7, '')
        _cell(r, 8, '')
        _cell(r, 9, '')

    # ── Cargo Handled section ────────────────────────────────────────────────
    cargo_start = 3 + len(ROWS)   # first row after ROWS (= row 21 currently)

    def _cargo_section(row_start, period_rows, period_label):
        r = row_start
        n = len(period_rows) + 1  # route rows + 1 total row
        # Column A: period label merged vertically over all rows
        _merge_col(r, r + n - 1, 1, period_label, bold=True, align=_ctr)
        # Route rows
        for route_name, qty in period_rows:
            _cell(r, 2, route_name, align=_left)
            _cell(r, 3, int(round(qty)) if qty else '', align=_ctr)
            for ci in range(4, 10):
                _cell(r, ci, '')
            ws.row_dimensions[r].height = 18
            r += 1
        # Total row
        total = sum(q for _, q in period_rows)
        _cell(r, 2, 'Total:', bold=True, align=_left)
        _cell(r, 3, int(round(total)) if total else '', bold=True, align=_ctr)
        for ci in range(4, 10):
            _cell(r, ci, '')
        ws.row_dimensions[r].height = 18
        r += 1
        return r

    if day_rows or month_rows:
        r = cargo_start
        # Blank separator row
        for ci in range(1, 10):
            _cell(r, ci, '')
        ws.row_dimensions[r].height = 18
        r += 1
        # "Cargo Handled" header — A:C merged, D:I bordered
        _merge_row(r, 1, 3, 'Cargo Handled', bold=True, align=_left)
        for ci in range(4, 10):
            _cell(r, ci, '')
        ws.row_dimensions[r].height = 18
        r += 1
        # For the Day
        if day_rows:
            r = _cargo_section(r, day_rows, 'For the Day')
        # For the Month
        if month_rows:
            r = _cargo_section(r, month_rows, 'For the Month')

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf


@bp.route('/api/module/RP01/daily-ops/download')
@login_required
def daily_ops_download():
    date_str       = request.args.get('report_date', date.today().strftime('%Y-%m-%d'))
    operation_type = request.args.get('operation_type', 'Import')

    if operation_type not in ('Import', 'Export'):
        return Response('Invalid operation type', status=400)
    try:
        report_date = datetime.strptime(date_str, '%Y-%m-%d').date()
    except ValueError:
        return Response('Invalid date', status=400)

    vessels = _fetch_data(report_date, operation_type)
    if not vessels:
        return Response(f'No active {operation_type} vessels on the selected date', status=404)

    day_rows, month_rows = _fetch_cargo_handled(report_date, operation_type)
    buf = _build_excel(vessels, report_date, operation_type, day_rows, month_rows)
    fname = f'DailyOps_{operation_type}_{date_str}.xlsx'
    return Response(
        buf.getvalue(),
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        headers={'Content-Disposition': f'attachment; filename="{fname}"'},
    )
